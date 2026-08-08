"""
Maintains data/postings.xlsx as a running history.

Each run of main.py creates a NEW SHEET (tab) inside this same workbook,
named with that run's date + timestamp (e.g. "2026-07-29_14-32-05").
Nothing from previous runs is ever overwritten or deleted -- full history
is preserved across every tab.

Deduplication still works across the ENTIRE file: before writing new rows,
we scan every existing sheet's posting_hash column and skip anything
already seen in any prior run, even though it now lives in a different tab.
"""

import os
import re
import hashlib
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

COLUMNS = [
    "posting_hash", "source", "company", "title", "location",
    "date_posted", "date_collected", "jd_text", "apply_url",
    "fit_score", "profile", "gate_status", "gate_reasons",
    "channel", "channel_kind", "gaps", "date_sent",
    "status", "applied_date", "follow_up_date",
    "response_date", "days_to_response", "outcome", "resume_file",
]

APPLY_URL_COL_INDEX = COLUMNS.index("apply_url") + 1


def _hash_posting(company, title, location="", apply_url=""):
    """Identity of a posting.

    BUG FIX: this used to hash company|title only. A company posting the SAME
    role in Berlin, Munich and Frankfurt collapsed into a single row -- you
    only ever saw one of the three. Location is now part of the identity.

    The apply_url is the strongest signal when present (it is unique per
    posting), so it wins. Query strings are stripped first so tracking
    parameters don't make the same posting look new on every run.
    """
    url = _strip_query(apply_url)
    if url:
        return hashlib.sha256(url.encode("utf-8")).hexdigest()[:16]

    norm = lambda x: " ".join((x or "").strip().lower().split())
    key = f"{norm(company)}|{norm(title)}|{norm(location)}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def _strip_query(url):
    """Drop ?tracking=... so the same posting hashes identically each run."""
    u = (url or "").strip()
    if not u or u.lower().startswith("n/a"):
        return ""
    return u.split("?")[0].split("#")[0].rstrip("/")


def _clean_url(url):
    if not url:
        return ""
    url = url.strip()
    if not url:
        return ""
    if not re.match(r"^https?://", url, re.IGNORECASE):
        if url.startswith("//"):
            url = "https:" + url
        elif re.match(r"^[\w.-]+\.[a-z]{2,}(/.*)?$", url, re.IGNORECASE):
            url = "https://" + url
        else:
            return ""
    if not re.match(r"^https?://[\w.-]+\.[a-z]{2,}", url, re.IGNORECASE):
        return ""
    return url


def _make_sheet_name():
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def load_all_existing_hashes(xlsx_path):
    if not os.path.exists(xlsx_path):
        return set()
    wb = load_workbook(xlsx_path)
    hashes = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                hashes.add(row[0])
    return hashes


def audit_history(xlsx_path):
    """Summarize the tracked history so you can confirm the dedupe net covers
    everything already applied to (the 'N already sent' question). Returns a
    dict of stats; also detects any duplicate hashes that should not exist."""
    if not os.path.exists(xlsx_path):
        return {"sheets": 0, "rows": 0, "unique": 0, "duplicate_rows": [],
                "applied": 0, "by_status": {}}
    wb = load_workbook(xlsx_path)
    rows = 0
    seen = set()
    dup_rows = []
    applied = 0
    by_status = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
        status_idx = headers.get("status")
        hash_idx = headers.get("posting_hash")
        applied_idx = headers.get("applied_date")
        for r in range(2, ws.max_row + 1):
            h = ws.cell(row=r, column=hash_idx).value if hash_idx else None
            if not h:
                continue
            rows += 1
            if h in seen:
                dup_rows.append((sheet_name, r))
            seen.add(h)
            status = ws.cell(row=r, column=status_idx).value if status_idx else None
            by_status[status or "unknown"] = by_status.get(status or "unknown", 0) + 1
            applied_date = ws.cell(row=r, column=applied_idx).value if applied_idx else None
            if applied_date:
                applied += 1
    return {"sheets": len(wb.sheetnames), "rows": rows, "unique": len(seen),
            "duplicate_rows": dup_rows, "applied": applied, "by_status": by_status}


def import_history(xlsx_path, csv_path):
    """Load a CSV of postings you already applied to into a dedicated sheet
    ('imported_history'), so the dedupe net covers them and the pipeline never
    re-packages a role you already burned.

    CSV columns: company,title,location,apply_url (header optional; apply_url
    may be blank). Rows whose hash already exists are skipped.
    """
    import csv
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        raw = list(reader)
    if not raw:
        return 0
    if raw[0] and raw[0][0].strip().lower() in ("company", "company_name"):
        header = raw[0]
        header = {name.strip().lower(): i for i, name in enumerate(header)}
        rows = raw[1:]
        pick = lambda row, *keys: next((row[i].strip() if i < len(row) else "" for i in [header.get(k) for k in keys] if i is not None), "")
    else:
        rows = raw
        pick = lambda row, *keys: (row[0].strip() if len(row) > 0 else "")

    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    existing = load_all_existing_hashes(xlsx_path)
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = "imported_history"
    counter = 1
    base = sheet_name
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base}_{counter}"
        counter += 1
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)

    added = 0
    from datetime import date
    today = date.today().isoformat()
    for row in rows:
        if not row or not row[0].strip():
            continue
        if header and all(not (row[i].strip() if i < len(row) else "") for i in header.values()):
            continue
        company = pick(row, "company", "company_name")
        title = pick(row, "title", "job_title", "role")
        location = pick(row, "location", "city")
        apply_url = pick(row, "apply_url", "url", "link")
        h = _hash_posting(company, title, location, apply_url)
        if h in existing:
            continue
        existing.add(h)
        clean_url = _clean_url(apply_url)
        ws.append([
            h, "history_import", company, title, location, "", today, "",
            clean_url if clean_url else "N/A - imported from history",
            "", "", "imported", "already applied (pre-existing history)",
            "", "", "", today, "applied", today, "", "", "", "", "",
        ])
        if clean_url:
            cell = ws.cell(row=ws.max_row, column=APPLY_URL_COL_INDEX)
            cell.hyperlink = clean_url
            cell.font = Font(color="0563C1", underline="single")
        added += 1

    if added:
        wb.save(xlsx_path)
    return added


def aggregate_gaps(xlsx_path):
    """Accumulate the per-posting `gaps` column across ALL sheets (R5).

    Returns a dict {gap_tech: {"count": n, "examples": [(company, title), ...]}}
    so the top blockers can be read empirically after many gated postings.
    """
    counts = {}
    if not os.path.exists(xlsx_path):
        return counts
    wb = load_workbook(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {c.value: i + 1 for i, c in enumerate(ws[1]) if c.value}
        gaps_idx = headers.get("gaps")
        company_idx = headers.get("company")
        title_idx = headers.get("title")
        if not gaps_idx:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = row[gaps_idx - 1] if len(row) >= gaps_idx else None
            if not raw:
                continue
            company = row[company_idx - 1] if company_idx and len(row) >= company_idx else ""
            title = row[title_idx - 1] if title_idx and len(row) >= title_idx else ""
            for gap in re.split(r"[,\n;]+", str(raw)):
                gap = gap.strip()
                if not gap:
                    continue
                entry = counts.setdefault(gap, {"count": 0, "examples": []})
                entry["count"] += 1
                if len(entry["examples"]) < 3:
                    entry["examples"].append((str(company), str(title)))
    return counts


def append_postings(xlsx_path, postings):
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    existing = load_all_existing_hashes(xlsx_path)

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

    sheet_name = _make_sheet_name()
    base_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name}_{counter}"
        counter += 1

    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)

    new_rows = []
    from datetime import date
    today = date.today().isoformat()

    for p in postings:
        h = _hash_posting(p["company"], p["title"],
                          p.get("location", ""), p.get("apply_url", ""))
        if h in existing:
            continue
        existing.add(h)

        clean_url = _clean_url(p.get("apply_url", ""))

        row = {
            "posting_hash": h,
            "source": p.get("source", ""),
            "company": p.get("company", ""),
            "title": p.get("title", ""),
            "location": p.get("location", ""),
            "date_posted": p.get("date_posted", ""),
            "date_collected": today,
            "jd_text": p.get("jd_text", ""),
            "apply_url": clean_url if clean_url else "N/A - broken or missing link, check source manually",
            "fit_score": p.get("fit_score", ""),
            "profile": p.get("profile", ""),
            "gate_status": p.get("gate_status", ""),
            "gate_reasons": p.get("gate_reasons", ""),
            "channel": p.get("submission_channel", ""),
            "channel_kind": p.get("submission_channel_kind", ""),
            "gaps": p.get("gaps", ""),
            "date_sent": p.get("date_sent", ""),
            "status": p.get("status", "not_applied"),
            "applied_date": p.get("applied_date", ""),
            "follow_up_date": p.get("follow_up_date", ""),
            "response_date": p.get("response_date", ""),
            "days_to_response": p.get("days_to_response", ""),
            "outcome": p.get("outcome", ""),
            "resume_file": p.get("resume_file", ""),
        }
        ws.append([row[c] for c in COLUMNS])

        if clean_url:
            cell = ws.cell(row=ws.max_row, column=APPLY_URL_COL_INDEX)
            cell.hyperlink = clean_url
            cell.font = Font(color="0563C1", underline="single")

        new_rows.append(row)

    wb.save(xlsx_path)
    print(f"New sheet '{sheet_name}' created with {len(new_rows)} new posting(s).")
    return new_rows