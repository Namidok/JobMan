"""
Follow-up engine: tracks which postings you actually applied to, when the
follow-up is due, and whether anything ever came back.

Generating a tailored CV+CL only gets you to "applied". Most internship offers
come from postings you follow up on, so this module turns the Excel log into a
living to-do list.

Usage (via main.py):
    python main.py --followup                        # print the full report
    python main.py --mark-applied SHEET ROW          # mark one posting as applied
    python main.py --mark-outcome SHEET ROW OUTCOME  # record replied/interview/offer/rejected/withdrawn

The report splits every logged posting into:
    overdue     -- applied, follow-up date passed, no outcome recorded
    waiting     -- applied, follow-up still in the future
    not_applied -- never marked as applied
    decided     -- an outcome has been recorded

Overdue rows print a ready-to-send follow-up email built ONLY from config.py
content (plus the same variant highlight used in the cover letter) -- nothing
invented.
"""

import os
import sys
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from config import CONTACT
from pipeline.excel_log import COLUMNS
from resume_builder.cover_letter import HIGHLIGHTS

FOLLOW_UP_DAYS = 7
OUTCOMES = {"replied", "interview", "offer", "rejected", "withdrawn"}

FOLLOWUP_TEMPLATE = """Dear Hiring Team at {company},

I applied for the {role} position recently and wanted to briefly follow up. I
remain very interested in the role -- {highlight}

I understand you're likely busy, so I appreciate you taking the time. I'm happy
to provide any additional information or a quick walkthrough of my work at your
convenience.

Best regards,
Srikar Kodi
{email} | {phone}
"""


def _col_map(ws):
    """Map header names -> column numbers, tolerating missing/renamed columns."""
    mapping = {}
    for cell in ws[1]:
        if cell.value:
            mapping[str(cell.value)] = cell.column
    return mapping


def _ensure_column(ws, col, name):
    """Return the column index for `name`, appending the header if absent
    (so the follow-up columns can be added to sheets created before this
    feature existed)."""
    idx = col.get(name)
    if idx:
        return idx
    idx = ws.max_column + 1
    ws.cell(row=1, column=idx).value = name
    col[name] = idx
    return idx


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        try:
            return date.fromisoformat(str(value)[:10])
        except ValueError:
            return None
    return None


def load_rows(xlsx_path):
    """Return a dict per data row across all sheets (old + new columns)."""
    if not os.path.exists(xlsx_path):
        return []
    wb = load_workbook(xlsx_path)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col = _col_map(ws)
        for r in range(2, ws.max_row + 1):
            def get(name):
                idx = col.get(name)
                return ws.cell(row=r, column=idx).value if idx else None

            posting_hash = get("posting_hash")
            if not posting_hash:
                continue
            rows.append({
                "sheet": sheet_name,
                "row": r,
                "posting_hash": posting_hash,
                "company": get("company"),
                "title": get("title"),
                "apply_url": get("apply_url"),
                "overlap_pct": get("overlap_pct"),
                "best_variant": get("best_variant"),
                "status": get("status"),
                "applied_date": get("applied_date"),
                "follow_up_date": get("follow_up_date"),
                "outcome": get("outcome"),
            })
    return rows


def mark_applied(xlsx_path, sheet, row):
    today = date.today()
    follow_up = today + timedelta(days=FOLLOW_UP_DAYS)
    wb = load_workbook(xlsx_path)
    ws = wb[sheet]
    col = _col_map(ws)

    updates = {
        "status": "applied",
        "applied_date": today.isoformat(),
        "follow_up_date": follow_up.isoformat(),
    }
    for name, val in updates.items():
        idx = _ensure_column(ws, col, name)
        ws.cell(row=int(row), column=idx).value = val

    wb.save(xlsx_path)
    print(f"Marked '{sheet}' row {row} as applied. Follow-up due: {follow_up.isoformat()} (+{FOLLOW_UP_DAYS}d).")


def mark_outcome(xlsx_path, sheet, row, outcome):
    outcome = outcome.strip().lower()
    if outcome not in OUTCOMES:
        raise SystemExit(
            f"OUTCOME must be one of: {', '.join(sorted(OUTCOMES))} (got '{outcome}')"
        )
    wb = load_workbook(xlsx_path)
    ws = wb[sheet]
    col = _col_map(ws)

    out_idx = _ensure_column(ws, col, "outcome")
    ws.cell(row=int(row), column=out_idx).value = outcome

    fu_idx = _ensure_column(ws, col, "follow_up_date")
    ws.cell(row=int(row), column=fu_idx).value = None

    wb.save(xlsx_path)
    print(f"Recorded outcome '{outcome}' for '{sheet}' row {row} (follow-up cleared).")


def build_report(rows):
    today = date.today()
    report = {"overdue": [], "waiting": [], "not_applied": [], "decided": []}
    for row in rows:
        outcome = (row["outcome"] or "").strip().lower()
        if outcome in OUTCOMES:
            report["decided"].append(row)
            continue
        status = (row["status"] or "").strip().lower()
        if status == "applied":
            fu = _to_date(row["follow_up_date"])
            if fu and fu < today:
                report["overdue"].append(row)
            else:
                report["waiting"].append(row)
        else:
            report["not_applied"].append(row)
    return report


def build_followup_email(row):
    variant = row["best_variant"] or "ai_ml"
    if variant not in HIGHLIGHTS:
        variant = "ai_ml"
    return FOLLOWUP_TEMPLATE.format(
        company=row["company"],
        role=row["title"],
        highlight=HIGHLIGHTS[variant].strip(),
        email=CONTACT["email"],
        phone=CONTACT["phone"],
    )


def _line(row):
    return (f"- {row['company']} | {row['title']} "
            f"(applied {row['applied_date']}, follow-up due {row['follow_up_date']})  "
            f"[{row['sheet']} row {row['row']}]  {row['apply_url'] or ''}")


def render_report(xlsx_path):
    rows = load_rows(xlsx_path)
    if not rows:
        print("No postings logged yet -- run main.py with a source first.")
        return

    report = build_report(rows)
    print(f"\nFollow-up report -- {date.today().isoformat()} -- {len(rows)} postings tracked\n")

    print(f"== {len(report['overdue'])} OVERDUE (follow up today) ==")
    if not report["overdue"]:
        print("  (none -- nice)")
    for r in sorted(report["overdue"], key=lambda x: (x["follow_up_date"] or "")):
        print(_line(r))
        print("  Follow-up email draft:")
        for line in build_followup_email(r).splitlines():
            print(f"    {line}")
        print()

    print(f"== {len(report['waiting'])} applied, follow-up not yet due ==")
    for r in sorted(report["waiting"], key=lambda x: (x["follow_up_date"] or "")):
        print(_line(r))

    print(f"\n== {len(report['not_applied'])} never applied ==")
    for r in sorted(report["not_applied"], key=lambda x: str(x["overlap_pct"] or 0), reverse=True):
        pct = r["overlap_pct"] if r["overlap_pct"] is not None else ""
        print(f"- {r['company']} | {r['title']} ({pct}% match)  "
              f"[{r['sheet']} row {r['row']}]  {r['apply_url'] or ''}")

    print(f"\n== {len(report['decided'])} decided ==")
    for r in sorted(report["decided"], key=lambda x: str(x["outcome"] or "")):
        print(f"- {r['company']} | {r['title']} -> {r['outcome']}")

    print("\nMark postings as you apply and update outcomes as replies come in:")
    print("  python main.py --mark-applied  SHEET ROW")
    print("  python main.py --mark-outcome  SHEET ROW replied|interview|offer|rejected|withdrawn")
